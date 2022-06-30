# -*- coding: utf-8 -*-
# @Time    : 2022/6/12 14:54
# @Author  : Jiaan Chen

import os
import sys
import glob
import h5py
import numpy as np
import torch
from os.path import join

os.environ["KMP_DUPLICATE_LIB_OK"] = "TRUE"
import sys

sys.path.append("..")
import cv2
from tqdm import tqdm
import openpyxl
from eval_run_tools import init_point_model, run_point_modelh5_2D, run_all_h5_3D
from eval_load_tools import get_data_and_label, get_3D_label


# test data
root_data_dir = 'F://DHP19EPC_dataset//test_MeanLabel_extract//'
h5datadir = 'F://DHP19EPC_dataset//test_MeanLabel//data//'
h5labeldir = 'F://DHP19EPC_dataset//test_MeanLabel//label//'

P_mat_dir = '../P_matrices/'
cameras_pos = np.load(join(P_mat_dir, 'camera_positions.npy'))

P_mat_cam1 = np.load(join(P_mat_dir, 'P4.npy'))
P_mat_cam2 = np.load(join(P_mat_dir, 'P1.npy'))
P_mat_cam3 = np.load(join(P_mat_dir, 'P3.npy'))
P_mat_cam4 = np.load(join(P_mat_dir, 'P2.npy'))
P_mats = [P_mat_cam1, P_mat_cam2, P_mat_cam3, P_mat_cam4]

cam_num = [4, 1, 3, 2]  # camera number list

cam1 = 2
cam2 = 3

# centers of the 2 used cameras
Point0 = (np.stack((cameras_pos[cam_num[cam1] - 1], cameras_pos[cam_num[cam2] - 1])))  # 3D label

exp_name = 'PointTrans'
model, args = init_point_model('PointTrans')

if not os.path.exists('table_results'):
    os.makedirs('table_results')

run_result = [[], []]  # 2D and 3D results

result_data = openpyxl.Workbook()
result_data.create_sheet(exp_name)
result_table = result_data.get_sheet_by_name(exp_name)

for sub in range(13, 18):
    mov_all = 0
    for session in range(1, 6):
        if session == 1:
            numMovements = 8
        elif session == 2:
            numMovements = 6
        elif session == 3:
            numMovements = 6
        elif session == 4:
            numMovements = 6
        elif session == 5:
            numMovements = 7

        for mov in range(1, numMovements + 1):

            if sub == 14 and session == 5 and mov == 3:
                mov_all += 1
                mpjpe3D_single_file = 0
                result_table.cell(mov_all, sub - 12, '{:.2f}'.format(mpjpe3D_single_file))
                continue

            mov_all += 1
            h5label_name = "S{}_session{}_mov{}_label.h5".format(sub, session, mov)
            h5label = h5py.File(join(h5labeldir, h5label_name), 'r')
            h5data_name = "S{}_session{}_mov{}.h5".format(sub, session, mov)
            h5data = h5py.File(join(h5datadir, h5data_name), 'r')

            print('\n **Running S{}_session{}_mov{}.h5** \n'.format(sub, session, mov))
            mpjpe3D_list = []
            pbar = tqdm(total=len(h5label['XYZ']) - 1)
            for frame in range(len(h5label['XYZ']) - 1):
                pbar.update(1)
                video_info = [sub, session, mov, frame]
                data_numpy_origin1, gt1, gt_mask1 = get_data_and_label(root_data_dir, sub, session, mov, frame, cam1)
                data_numpy_origin2, gt2, gt_mask2 = get_data_and_label(root_data_dir, sub, session, mov, frame, cam2)

                data = [[data_numpy_origin1, gt1, gt_mask1], [data_numpy_origin2, gt2, gt_mask2]]
                label_3d = get_3D_label(root_data_dir, sub, session, mov, frame)

                Point_2Dresult = run_point_modelh5_2D(data, model, args)

                Point_3Dresult = run_all_h5_3D(Point_2Dresult[0], Point_2Dresult[1], P_mats, Point0, label_3d.T)

                run_result[0].append(Point_2Dresult[2])
                run_result[0].append(Point_2Dresult[3])
                run_result[1].append(Point_3Dresult[1])
                mpjpe3D_list.append(Point_3Dresult[1])

            mpjpe3D_single_file = np.nanmean(mpjpe3D_list)
            pbar.close()
            result_table.cell(mov_all, sub - 12, '{:.2f}'.format(mpjpe3D_single_file))

result_data.save('./table_results/' + exp_name + '.xlsx')

print(exp_name + ' result ==> 2D:{:.6f}, 3D:{:.6f}'.format(np.mean(run_result[0]), np.mean(run_result[1])))
print('******** Finish eval all files ********')
